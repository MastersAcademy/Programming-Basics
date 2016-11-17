#Homework 2

print ("Hi! Tell me something about yourself")

# Collecting of all answers into one dictionary
answers = {}

answers['name'] = input("What is your name?")
answers['age'] = int(input("How old are you?"))
answers['city'] = input("Where do you live?")
answers['movies'] = input("Do you like watch movies?")
if answers['movies'] == 'yes':
    #Collecting answer as list
    answers['what_movies'] = input("What kind of movies you like? (You can enter a few genres, separate them by comma)").split(',')

print(answers)

# Saving answers into .xslx file using openpyxl
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws['A1'] = "Name"
ws['A2'] = answers['name']
ws['B1'] = "Age"
ws['B2'] = answers['age']
ws['C1'] = "City"
ws['C2'] = answers['city']
ws['D1'] = "Watch movies"
ws['D2'] = answers['movies']
ws['E1'] = "What movies"
ws['E2'] = ','.join(map(str,answers['what_movies']))
wb.save("answer.xlsx")