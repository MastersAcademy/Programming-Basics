#Homework 1

print ("Hi! Tell me something about yourself")

# Questions
name = input("What is your name?")
age = int(input("How old are you?"))
city = input("Where do you live?")
movies = input("Do you like watch movies?")
if movies == 'yes':
    what_movies = input("What kind of movies you like?")
    print("So your name is %s, your age is %i, you live in %s and you like watch %s movies" %(name, age, city, what_movies))

if movies == 'no':
    print ("So your name is %s, your age is %i, you live in %s and you don't like watch movies" %(name, age, city))

# Saving answers into .xslx file using openpyxl
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws['A1'] = "Name"
ws['A2'] = "%s" %(name)
ws['B1'] = "Age"
ws['B2'] = "%i" % (age)
ws['C1'] = "City"
ws['C2'] = "%s" %(city)
ws['D1'] = "Watch movies"
ws['D2'] = "%s" %(movies)
ws['E1'] = "What movies"
ws['E2'] = "%s" %(what_movies)
wb.save("answer.xlsx")